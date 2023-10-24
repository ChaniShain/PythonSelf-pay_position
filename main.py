import cv2
import numpy as np
import pandas as pd
import pyzbar
from pyzbar.pyzbar import decode
import tkinter as tk
from tkinter import simpledialog
import openpyxl

wb = openpyxl.load_workbook('products.xlsx')
sheet = wb.active
product_info = {}
#קריאת הנתונים של המוצרים בחנות ופרטיהם מקובץ XL
for i in range(2,sheet.max_row+1):
    barcode=str(sheet.cell(i,3).value)
    product_name=sheet.cell(i,1).value
    price = float(sheet.cell(i,2).value)
# product_infoהכנסת הפרטים ל
    product_info[barcode] = {'name': product_name, 'price': price}
#קובץ שלתוכו יכתבו פרטי הקניות
try:
    orders_df = pd.read_csv('orders.csv')
except FileNotFoundError:
    orders_df = pd.DataFrame(columns=['Product Name', 'Quantity', 'Price'])
#  המצלמה
cap = cv2.VideoCapture(0)
# הגדרת SET כדי שלא יעבירו מוצר פעמיים
registered_barcodes = set()

root = tk.Tk()
root.geometry("800x200")
root.withdraw()

order = []
total_price = 0
# פונקציה שמסימת את ההרצה כשלוחצים על כפתור סיום הקניה
def button_callback(event, x, y, flags, param):
    if event == cv2.EVENT_LBUTTONDOWN:
        button_x, button_y, button_w, button_h = param
        if button_x <= x < button_x + button_w and button_y <= y < button_y + button_h:
            cap.release()
            cv2.destroyAllWindows()

cv2.namedWindow('Frame')


cv2.setMouseCallback('Frame', button_callback)
# הפעלת המצלמה
while True:
    ret, frame = cap.read()

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    # קריאה לpyzbar  שמפענחת ברקודים
    barcodes = pyzbar.pyzbar.decode(gray)
    # הגדרת טקסט וכפתור שיופיעו על החלונית בריצה
    text = 'Please scan the product barcode'
    cv2.putText(frame, text, (50, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0), 2)
    button_text = 'finish shopping'
    button_size = cv2.getTextSize(button_text, cv2.FONT_HERSHEY_SIMPLEX, 1, 2)[0]
    button_x = int((frame.shape[1] - button_size[0]) / 2)
    button_y = int(frame.shape[0] * 0.8)
    button_w = button_size[0] + 20
    button_h = button_size[1] + 20
    cv2.rectangle(frame, (button_x, button_y), (button_x + button_w, button_y + button_h), (255, 255, 255), -1)
    cv2.putText(frame, button_text, (button_x + 10, button_y + button_h - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0),2)
    cv2.setMouseCallback('Frame', button_callback, param=(button_x, button_y, button_w, button_h))

    # פיענוח הברקודים וזיהויים
    for barcode in barcodes:
        barcode_data = barcode.data.decode("utf-8")
        barcode_type = barcode.type

        # אם המוצר עוד לא קיים בקנייה
        if barcode_data not in registered_barcodes:
            registered_barcodes.add(barcode_data)
            (x, y, w, h) = barcode.rect
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            print("Found {} barcode: {}".format(barcode_type, barcode_data))

            #  בדיקה אם הברקוד קים ברשימת המוצרים בחנות
            if barcode_data in product_info:
                # אם כן שולפים את המידע המתאים של המוצר
                product_name = product_info[barcode_data]['name']
                price = product_info[barcode_data]['price']
                # המשתמש מקיש את כמות היחידות שהוא רוצה מהמוצר
                quantity = simpledialog.askinteger(title="כמות מוצר ", prompt="הקש מספר יחידות  {}\t\t\t".format(product_name))

                # מוסיפים את המוצר ופרטיו למערך order
                order.append({'Product Name': product_name, 'Quantity': quantity, 'Price': price})
                print(quantity)
                print(price)
                # עדכון הסכום הסופי של הקנייה
                total_price += float(quantity) * float(price)

                # שומרים את הנתונים בקובץ order
                orders_df = orders_df.append({'Product Name': product_name, 'Quantity': quantity, 'Price': price}, ignore_index=True)
                orders_df.to_csv('orders.csv', index=False)
             # אם הברקוד לא קים במוצרים של המערכת
            else:
                print('Barcode not found in product information')

    cv2.imshow("Frame", frame)

    # סיום התוכנית בתנאים המתאימים...
    if cv2.waitKey(1) & 0xFF == ord('q') :
        break
    if cv2.getWindowProperty('Frame', cv2.WND_PROP_VISIBLE) < 1:
        break

        if cv2.getWindowProperty('Frame', cv2.WND_PROP_AUTOSIZE) < 1:
            break
        if cv2.getWindowProperty('Frame', cv2.WND_PROP_FULLSCREEN) == cv2.WINDOW_FULLSCREEN:
            break
        if cv2.getWindowProperty('Frame', cv2.WND_PROP_FULLSCREEN) == cv2.WINDOW_NORMAL:
            if cv2.waitKey(1) & 0xFF == ord('c'):
                x, y = np.random.randint(0, frame.shape[1]), np.random.randint(0, frame.shape[0])
                cv2.circle(frame, (x, y), 50, (0, 0, 255), -1)

cap.release()
cv2.destroyAllWindows()

# הדפסת נתוני הקניה למשתמש:
print('Order Information:')
i=0
root2 = tk.Tk()

e1 = tk.Entry(root2)
e2 = tk.Entry(root2)
e3 = tk.Entry(root2)
e4 = tk.Entry(root2)
e5 = tk.Entry(root2)
e6 = tk.Entry(root2)
e1.grid(row=i, column=0)
e2.grid(row=i, column=1)
e3.grid(row=i, column=2)
e4.grid(row=i, column=3)
e4.insert(0,"מוצר")
e3.insert(0,"מחיר יחידה")
e2.insert(0,"כמות")
e1.insert(0,"מחיר")

i = 1
for item in order:
    print('{} x {} - ${}'.format(item['Product Name'], item['Quantity'], item['Price']))

    e1 = tk.Entry(root2)
    e2 = tk.Entry(root2)
    e3 = tk.Entry(root2)
    e4 = tk.Entry(root2)

    e1.grid(row=i, column=0)
    e2.grid(row=i, column=1)
    e3.grid(row=i, column=2)
    e4.grid(row=i, column=3)

    e4.insert(0, item['Product Name'])
    e3.insert(0, item['Price'])
    e2.insert(0, item['Quantity'])
    e1.insert(0, item['Price']*item['Quantity'])
    i = i + 1

e5 = tk.Entry(root2)
e6 = tk.Entry(root2)

e6.grid(row=i+1, column=2, rowspan=2, sticky='NESW')
e5.grid(row=i+1, column=1, rowspan=2, sticky='NESW')
e6.insert(0, ":לתשלום ")
e5.insert(0, total_price )

# בדיקה אם מספר האשראי שנכנס תקין
def check(num):
    while len(num) < 8 or len(num) > 8:
        return
    root2.quit()
    cap.release()
    cv2.destroyAllWindows()

# חלונית שלתוכה מקישים את מספר האשראי
def card():
    root3 = tk.Tk()
    root3.geometry("160x100")
    e7 = tk.Entry(root3)
    e8 = tk.Entry(root3)
    e7.grid(row=0, column=0, columnspan=2, sticky='NESW')
    e8.grid(row=1, column=0, columnspan=2, sticky='NESW')
    e7.insert(0," הכנס מספר אשראי 8 ספרות")
    e8.insert(0,"")
    btn_2 = tk.Button(root3, text="לסיום", command=lambda: check(e8.get()))
    btn_2.grid(row=2, column=0,columnspan=2, sticky='NESW')

def end_program():
    if total_price==0:
        cap.release()
        cv2.destroyAllWindows()
        root2.quit()
    else:
        card()


btn_0 = tk.Button(root2, text="לסיום",command=lambda: end_program())
btn_0.grid(row=i+3,column=1,sticky='NESW')
root2.mainloop()



